#
# from openpyxl import load_workbook
# import re
# import requests
# from bs4 import BeautifulSoup
#
#
# companyList = []
# ignoreText1 = "FiscalYear 2018"
# ignoreText2 = "Supplier Name (supplier number)"
#
#
# wb = load_workbook(filename='DSLB.xlsx', read_only=True)
# ws = wb['Sheet1']
#
# for row in ws.rows:
#     for cell in row:
#         cv = cell.value
#         if (cv is not None and cv != ignoreText1 and cv != ignoreText2):
#             cv = re.sub("\([^>]+\)", "", cv)
#             # print(cv)
#             companyList.append(cv)
#
#
#
#
# def companyTotalResults(theCompanyName):
#     # companyName = "Dicks Sporting Goods"
#     companyName = theCompanyName
#
#     companyNamePlus = re.sub(" ", "+", companyName)
#     companyNamePlus = re.sub("\n", "", companyNamePlus)
#     companyNamePlus = companyNamePlus.lower()
#
#     # print(companyNamePlus)
#
#     myLink = "https://www.google.com/search?q=" + "\"" + companyNamePlus + "\"" + "+\"prison+labor\""
#     page = requests.get(myLink)
#
#     soup = BeautifulSoup(page.content, 'html.parser')
#
#     container = soup.find(id="resultStats")
#     tempCount = "" + str(container)
#
#     # print(tempCount)
#
#     tempCount = tempCount.replace("<div class=\"sd\" id=\"resultStats\">About ", "")
#     tempCount = tempCount.replace(" results</div>", "")
#     tempCount = tempCount.replace(",", "")
#
#     result = 0
#     try:
#         result = int(tempCount)
#     except:
#         result = 0
#
#     return result
#
#
#
# companyCountList = []
#
# for index in range(len(companyList)):
#     # print(companyList[index] + " " + str(companyTotalResults(companyList[index])))
#     companyCountList.append(companyTotalResults(companyList[index]))
#
#
#
#
# for i in range(len(companyList)):
#     biggestCount = 0
#     for k in range(len(companyList)):
#         if (int(companyCountList[k]) > biggestCount):
#             biggestCount = companyCountList[k]
#
#     tempIndex = companyCountList.index(biggestCount)
#     print(companyList[tempIndex] + " " + str(companyCountList[tempIndex]))
#
#     companyList.remove(companyList[tempIndex])
#     companyCountList.remove(companyCountList[tempIndex])
#
# ############################################

# GUI

fileLocation = ""

from tkinter import filedialog
from tkinter import messagebox
from tkinter import *
import xlsxwriter
import os, sys



# Create the window
root = Tk()
v = IntVar()


def loadFile():
    fileLocation = filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("excel files","*.xlsx"),))
    if (fileLocation != ""):
        # run search
        from openpyxl import load_workbook
        import re
        import requests
        from bs4 import BeautifulSoup


        companyList = []
        ignoreText1 = "FiscalYear 2018"
        ignoreText2 = "Supplier Name (supplier number)"


        wb = load_workbook(filename=fileLocation, read_only=True)
        ws = wb['Sheet1']

        for row in ws.rows:
            for cell in row:
                cv = cell.value
                if (cv is not None and cv != ignoreText1 and cv != ignoreText2):
                    cv = re.sub("\([^>]+\)", "", cv)
                    # print(cv)
                    companyList.append(cv)




        def companyTotalResults(theCompanyName):
            # companyName = "Dicks Sporting Goods"
            companyName = theCompanyName

            companyNamePlus = re.sub(" ", "+", companyName)
            companyNamePlus = re.sub("\n", "", companyNamePlus)
            companyNamePlus = companyNamePlus.lower()

            # print(companyNamePlus)

            myLink = "https://www.google.com/search?q=" + "\"" + companyNamePlus + "\"" + "+\"prison+labor\""
            page = requests.get(myLink)

            soup = BeautifulSoup(page.content, 'html.parser')

            container = soup.find(id="resultStats")
            tempCount = "" + str(container)

            # print(tempCount)

            tempCount = tempCount.replace("<div class=\"sd\" id=\"resultStats\">About ", "")
            tempCount = tempCount.replace(" results</div>", "")
            tempCount = tempCount.replace(",", "")

            result = 0
            try:
                result = int(tempCount)
            except:
                result = 0

            return result



        companyCountList = []

        for index in range(len(companyList)):
            # print(companyList[index] + " " + str(companyTotalResults(companyList[index])))
            companyCountList.append(companyTotalResults(companyList[index]))

        # Create an new Excel file and add a worksheet.
        workbook = xlsxwriter.Workbook('Company Prison Labor Search.xlsx')
        worksheet = workbook.add_worksheet()

        for i in range(len(companyList)):
            biggestCount = 0
            for k in range(len(companyList)):
                if (int(companyCountList[k]) > biggestCount):
                    biggestCount = companyCountList[k]

            tempIndex = companyCountList.index(biggestCount)
            ##################
            # print(companyList[tempIndex] + " " + str(companyCountList[tempIndex]))
            worksheet.write(i, 0, companyList[tempIndex])
            worksheet.write(i, 1, str(companyCountList[tempIndex]))
            ##################

            companyList.remove(companyList[tempIndex])
            companyCountList.remove(companyCountList[tempIndex])

        workbook.close()


        # msg = messagebox.showinfo( "Done!", "Please open \"Company Prison Labor Search.xlsx\" to view result.")
        # fd = os.open( "Company Prison Labor Search.xlsx")
        try:
            os.system('open Company\ Prison\ Labor\ Search.xlsx')
        except:
            fd = os.open( "Company Prison Labor Search.xlsx")

    else:
        msg = messagebox.showinfo( "Error", "File did not load properly")






#modify root window
root.title("Company Search Tool")
root.geometry("200x50")

app = Frame(root)
app.grid()
label = Label(app, text = "Company Search:")
selectButton = Button(app, text = "Select File", command = loadFile)


label.grid()
selectButton.grid()

root.mainloop()





















#
